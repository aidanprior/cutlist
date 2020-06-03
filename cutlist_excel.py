from openpyxl import load_workbook, Workbook
import openpyxl
from sys import argv
from cutlist import Cutlist, Length, Board
from data_classes import Angle


        # print(type(quantities[i]), type(lengths[i]), type(numerators[i]), type(a1), type(a2), type(ol1), type(ol2))

""" csv or worksheet should be formatted in the following way:
stock, inches, 64ths, x, inches 64ths (where the first length is the width, and the second the thickness)
in row 1 with the following table below it:
quantity, left angle, right angle, length, 64ths  (where and angle is formatted: "45:W" or "45:T")
as columns w/ headers in A,B,C,D,E on the BOM sheet """
def generate_empty_dom():
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    A = ["(Optional)", "Project Name:"]
    B = ["(Optional)", "Part Name:"]
    C = ["(Required)", "Quantity:"]
    D = ["(Required)", "Stock:"]
    E = ["(Required)", "Length:"]
    F = ["(Optional)", "Angle 1:"]
    G = ["(Optional)", "Angle 2:"]
    H = []
    I = ["Angle Ex.", "", "Stock Ex.", "", "Length Ex."]
    J = ["45:W", "", "2x4", "", "4'-7 1/2\""]
    K = ["NOTE: On a 2x4, this would mean cutting a 45 degree angle along the width or 4 dimenstion, the other option is :T for thickness",
    "NOTE: Angled boards should have a length equal to the linear distance from farthest tip to farthest tip",
    'NOTE: Thickness followed by width, so lead with the smaller number. Don’t include "s',
    "NOTE: Length can be in ft'-in\", just inches (followed by \"), include fractional inches (with a space between the fraction and whole, and followed by \"), or a decimal"
    ]
    columns = [A, B, C, D, E, F, G, H, I, J, K]

    for column_number, column in enumerate(columns):
        row = 1
        for cell_value in column:
            ws.cell(row=row, column=column_number+1, value=cell_value)
            row += 1

    file_path = "C:\\Users\\AWP95\\Desktop\\Empty BOM.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path

def load_boards(wb):
    ws = wb['BOM']
    boards = {}
    project_names, part_names, quantities, stocks, lengths, angle_1s, angle_2s = ws.iter_cols(min_row=3, max_col=7, values_only=True)
    for i, qty in enumerate(quantities):
        if qty == None: #hit the bottom of the list
            break
        
        if stocks[i] not in boards:
            boards[stocks[i]] = []
        else:
            stock_width = Length(stocks[i].split('x')[1])
            stock_thickness = Length(stocks[i].split('x')[0])

        angle1 = None
        if angle_1s[i]: #angle 1 specified
            a = float(angle_1s[i].split(':')[0])
            ol = stock_width if angle_1s[i].split(':')[1] == "W" else stock_thickness
            angle1 = Angle(a, ol)
        
        angle2 = None
        if angle_2s[i]: #angle 2 specified
            a = float(angle_2s[i].split(':')[0])
            ol = stock_width if angle_2s[i].split(':')[1] == "W" else stock_thickness
            angle2 = Angle(a, ol)

        for j in range(qty):
            boards[stocks[i]].append(Board(project_names[i], part_names[i], Length(lengths[i]), angle1, angle2))
    return boards


def write_cut_list(wb, stock, quantity_wastes_used_boards_lists, total_waste, needed_stock_boards):
    ws = wb.create_sheet(stock + "s")
    ws['A1'] = needed_stock_boards
    ws['B1'] = stock + "s"
    ws['D1'] = "Waste:"
    ws['E1'] = str(total_waste)

    start_row = 1
    max_row = 1
    for index, (quantity, waste, used_boards) in enumerate(quantity_wastes_used_boards_lists):
        start_column = 4*(index%3) + 1
        if start_column == 1:
            start_row = max_row + 1
            max_row += 1
        
        print("Writing used boards list #" + str(index+1))
        row = write_used_boards(ws, quantity, waste, used_boards, start_row, start_column)
        max_row = row if row > max_row else max_row

    #can't find good documentation on this stuff, so we just try everything
    ws.page_setup.scale = ws.page_setup.fitToWidth
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_area = ws.calculate_dimension()
    ws.orientation = ws.ws.ORIENTATION_LANDSCAPE
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.print_options.gridLines = True
    
    ws.orientation = ws.ORIENTATION_LANDSCAPE
        

def write_used_boards(ws, quantity, waste, used_boards, row, column):
    ws.cell(row=row, column=column, value= quantity)
    ws.cell(row=row, column=column+1, value= "Waste:")
    ws.cell(row=row, column=column+2, value= str(waste))
    ws.cell(row=row+1, column=column, value= "Length:")
    ws.cell(row=row+1, column=column+1, value= "Angle 1:")
    ws.cell(row=row+1, column=column+2, value= "Angle 1:")

    row += 2
    for board in used_boards:
        ws.cell(row=row, column=column, value= str(board.length))
        ws.cell(row=row, column=column+1, value= str(board.angle1))
        ws.cell(row=row, column=column+2, value= str(board.angle2))
        row += 1
    return row

if __name__ == "__main__":
    if len(argv) > 2:
        print("Error: Too many arguments passed!")
        print("Exiting...")
    elif len(argv) < 2:
        print("Generating empty BOM spreadsheet...")
        print("Saved to: '", generate_empty_dom(), "'")
        print("Exiting...")
    else:
        print("Loading workbook " + argv[1] + "...")
        wb_in = load_workbook(argv[1])
        wb_out = Workbook()
        delete = wb_out.active
        print("Loading boards...")
        boards = load_boards(wb_in)

        for stock in boards:
            boards_list = boards[stock]
            print("Generating cut list for " + stock + "s... ")
            cutlist = Cutlist(boards_list, Length(120, 0), Length(0, 4))
            needed_stock_boards, total_waste, quantity_wastes_used_boards_lists = cutlist.solve()
            print("Writing cut list for " + stock + "s...")
            write_cut_list(wb_out, stock,quantity_wastes_used_boards_lists, total_waste, needed_stock_boards)
        
        print("Saving...")
        wb_out.remove(delete)
        wb_out.save(argv[1].split("BOM")[0] + "Cut List.xlsx")
        wb_in.close()
        wb_out.close()
        print("Exiting...")